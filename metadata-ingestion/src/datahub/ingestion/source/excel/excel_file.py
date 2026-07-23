import logging
import re
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Dict, Iterator, List, Optional, Union

import openpyxl
import pandas as pd
from openpyxl.workbook import Workbook

from datahub.ingestion.source.excel.report import ExcelSourceReport

logger = logging.getLogger(__name__)


@dataclass
class ExcelTable:
    df: pd.DataFrame
    header_row: int
    footer_row: int
    row_count: int
    column_count: int
    metadata: Dict[str, Any]
    sheet_name: str


class ExcelFile:
    wb: Workbook
    filename: str
    data: BytesIO
    sheet_list: List[str]
    active_sheet: str
    properties: Dict[str, Any]
    report: ExcelSourceReport

    def __init__(
        self,
        filename: str,
        data: BytesIO,
        report: ExcelSourceReport,
    ) -> None:
        self.filename = filename
        self.data = data
        self.report = report
        self.sheet_list = []
        self.active_sheet = ""
        self.properties = {}

    def load_workbook(self) -> bool:
        try:
            self.wb = openpyxl.load_workbook(self.data, data_only=True)
            self.properties = self.read_excel_properties(self.wb)
            self.sheet_list = self.wb.sheetnames
            self.active_sheet = self.wb.active.title
            return True
        except Exception as e:
            self.report.report_file_dropped(self.filename)
            self.report.warning(
                message="Error reading Excel file",
                context=f"Filename={self.filename}",
                exc=e,
            )
            return False

    @property
    def sheet_names(self) -> List[str]:
        return self.sheet_list

    @property
    def active_sheet_name(self) -> str:
        return self.active_sheet

    @property
    def workbook_properties(self) -> Dict[str, Any]:
        return self.properties

    def get_tables(self, active_only: Optional[bool] = False) -> Iterator[ExcelTable]:
        sheet_list = [self.active_sheet] if active_only else self.sheet_list
        for sheet in sheet_list:
            table = self.get_table(sheet)
            if table is not None:
                yield table
            else:
                self.report.report_worksheet_dropped(sheet)
                self.report.warning(
                    message="Worksheet does not contain a table",
                    context=f"Worksheet=[{self.filename}]{sheet}",
                )

    def get_table(self, sheet_name: str) -> Union[ExcelTable, None]:
        sheet = self.wb[sheet_name]

        # Extract all rows from the sheet
        rows = [[cell.value for cell in row] for row in sheet.rows]

        # Find a potential header row
        header_row_idx = self.find_header_row(rows)
        if header_row_idx is None:
            return None

        # Find where the footer starts
        footer_start_idx = self.find_footer_start(rows, header_row_idx)

        # Extract metadata before the header
        header_metadata = self.extract_metadata(rows[:header_row_idx])

        # Extract footer metadata
        footer_metadata = {}
        if footer_start_idx < len(rows):
            footer_metadata = self.extract_metadata(rows[footer_start_idx:])

        # Combine metadata
        metadata = {}
        metadata.update(self.properties)

        # Add header metadata
        for key, value in header_metadata.items():
            if key not in metadata:
                metadata[key] = value
            else:
                metadata[f"{key}_1"] = value

        # Add footer metadata
        for key, value in footer_metadata.items():
            if key not in metadata:
                metadata[key] = value
            else:
                metadata[f"{key}_1"] = value

        # Get the header row
        header_row = rows[header_row_idx]

        # Find the last non-empty column in the header row
        last_non_empty_idx = -1
        for i in range(len(header_row) - 1, -1, -1):
            if header_row[i] is not None and str(header_row[i]).strip() != "":
                last_non_empty_idx = i
                break

        # Truncate the header row to remove empty trailing columns
        if last_non_empty_idx >= 0:
            header_row = header_row[: last_non_empty_idx + 1]

        # Create the column names for the DataFrame
        column_names: List[str] = []
        seen_columns: Dict[str, int] = {}
        for i, col in enumerate(header_row):
            if col is None or str(col).strip() == "":
                col_name = f"Unnamed_{i}"
            else:
                col_name = str(col).strip()

            if col_name in seen_columns:
                seen_columns[col_name] += 1
                col_name = f"{col_name}_{seen_columns[col_name]}"
            else:
                seen_columns[col_name] = 0

            column_names.append(col_name)

        # Create the DataFrame with the table data
        data_rows = rows[header_row_idx + 1 : footer_start_idx]

        # Truncate data rows to match the header length
        truncated_data_rows = [
            row[: len(column_names)] if len(row) > len(column_names) else row
            for row in data_rows
        ]

        # Create the final DataFrame
        df = pd.DataFrame(truncated_data_rows, columns=column_names)

        row_count = df.shape[0]
        column_count = df.shape[1]

        return ExcelTable(
            df,
            header_row_idx + 1,
            footer_start_idx,
            row_count,
            column_count,
            metadata,
            sheet.title.strip(),
        )

    def find_header_row(self, rows: List[List[Any]]) -> Union[int, None]:
        max_score = -1
        header_idx = 0

        # Skip empty rows at the beginning
        start_idx = self._find_first_non_empty_row(rows)

        # Evaluate each potential header row with a lookahead
        min_rows_required = 2

        # Skip evaluation if there aren't enough rows
        if len(rows) < start_idx + min_rows_required + 1:
            return header_idx

        for i in range(start_idx, len(rows) - min_rows_required):
            current_row = rows[i]
            # Take as many next rows as available, up to 3
            next_rows = rows[i + 1 : min(i + 4, len(rows))]

            # Skip empty rows
            if not self._is_non_empty_row(current_row):
                continue

            score = self._calculate_header_row_score(i, current_row, next_rows, rows)

            if score > max_score:
                max_score = score
                header_idx = i

        if max_score <= 0:
            return None
        else:
            return header_idx

    def _find_first_non_empty_row(self, rows: List[List[Any]]) -> int:
        for i, row in enumerate(rows):
            if self._is_non_empty_row(row):
                return i
        return 0

    @staticmethod
    def _is_non_empty_row(row: List[Any]) -> bool:
        return any(cell is not None and str(cell).strip() != "" for cell in row)

    def _calculate_header_row_score(
        self,
        row_idx: int,
        current_row: List[Any],
        next_rows: List[List[Any]],
        all_rows: List[List[Any]],
    ) -> int:
        score = 0

        score += ExcelFile._score_row_with_numeric_cells(current_row)
        if score < 0:
            return score
        score += self._score_non_empty_cells(row_idx, current_row, all_rows)
        score += self._score_header_like_text(current_row)
        score += self._score_text_followed_by_numeric(current_row, next_rows)
        score += self._score_column_type_consistency(current_row, next_rows)
        score += self._score_metadata_patterns(row_idx, current_row, all_rows)

        return score

    @staticmethod
    def _score_non_empty_cells(
        row_idx: int, current_row: List[Any], all_rows: List[List[Any]]
    ) -> int:
        if row_idx <= 0:
            return 0

        non_empty_current = sum(
            1 for cell in current_row if cell is not None and str(cell).strip() != ""
        )

        non_empty_prev = sum(
            1
            for cell in all_rows[row_idx - 1]
            if cell is not None and str(cell).strip() != ""
        )

        return 2 if non_empty_current > non_empty_prev else 0

    @staticmethod
    def _score_header_like_text(row: List[Any]) -> int:
        return sum(
            1
            for cell in row
            if cell is not None
            and isinstance(cell, str)
            and re.match(r"^[A-Z][a-zA-Z\s]*$", str(cell).strip())
        )

    @staticmethod
    def _score_row_with_numeric_cells(row: List[Any]) -> int:
        return sum(
            -1 for cell in row if cell is not None and isinstance(cell, (int, float))
        )

    @staticmethod
    def _score_text_followed_by_numeric(
        current_row: List[Any], next_rows: List[List[Any]]
    ) -> int:
        if not next_rows:
            return 0

        header_text_count = sum(
            1 for cell in current_row if cell is not None and isinstance(cell, str)
        )

        next_rows_numeric_count = [
            sum(
                1
                for cell in row
                if cell is not None
                and (
                    isinstance(cell, (int, float))
                    or (
                        isinstance(cell, str)
                        and re.match(r"^-?\d+(\.\d+)?$", str(cell).strip())
                    )
                )
            )
            for row in next_rows
        ]

        if header_text_count > 0 and any(
            count > 0 for count in next_rows_numeric_count
        ):
            return 6 + sum(1 for count in next_rows_numeric_count if count > 0)
        return 0

    @staticmethod
    def _score_column_type_consistency(
        current_row: List[Any], next_rows: List[List[Any]]
    ) -> int:
        if len(next_rows) < 2:
            return 0

        col_types = []
        for col_idx in range(len(current_row)):
            if col_idx < len(current_row) and current_row[col_idx] is not None:
                col_type_counter: Counter = Counter()
                for row in next_rows:
                    if col_idx < len(row) and row[col_idx] is not None:
                        cell_type = type(row[col_idx]).__name__
                        col_type_counter[cell_type] += 1

                if col_type_counter and col_type_counter.most_common(1)[0][1] >= 1:
                    col_types.append(col_type_counter.most_common(1)[0][0])

        return 3 if len(col_types) >= 2 and len(set(col_types)) >= 1 else 0

    @staticmethod
    def _score_metadata_patterns(
        row_idx: int, current_row: List[Any], all_rows: List[List[Any]]
    ) -> int:
        score = 0

        if row_idx == 0 and len(current_row) <= 2:
            metadata_like = sum(
                1
                for cell in current_row
                if cell is not None and isinstance(cell, str) and len(str(cell)) <= 20
            )
            if metadata_like <= 2:
                score -= 1

        if row_idx < len(all_rows) - 1 and len(current_row) >= 2:
            if all(
                isinstance(cell, str) for cell in current_row[:2] if cell is not None
            ):
                score -= 2

        return score

    @staticmethod
    def find_footer_start(rows: List[List[Any]], header_row_idx: int) -> int:
        if header_row_idx + 1 >= len(rows):
            return len(rows)

        # Start with the assumption that all rows after the header are data (no footer)
        footer_start_idx = len(rows)

        # Get the number of columns in the header row to determine table width
        header_row = rows[header_row_idx]
        table_width = sum(
            1 for cell in header_row if cell is not None and str(cell).strip() != ""
        )

        # Get a sample of data rows to establish patterns
        data_sample_idx = min(header_row_idx + 5, len(rows) - 1)
        data_rows = rows[header_row_idx + 1 : data_sample_idx + 1]

        # Check for rows with significantly fewer populated cells than the data rows
        avg_populated_cells = sum(
            sum(1 for cell in row if cell is not None and str(cell).strip() != "")
            for row in data_rows
        ) / len(data_rows)

        # Look for pattern breaks, empty rows, or format changes
        for i in range(header_row_idx + 1, len(rows)):
            current_row = rows[i]

            # Skip completely empty rows unless followed by non-data-like rows
            if not any(
                cell is not None and str(cell).strip() != "" for cell in current_row
            ):
                # Look ahead to see if this empty row marks the start of footer
                if i + 1 < len(rows):
                    next_row = rows[i + 1]
                    next_row_populated = sum(
                        1
                        for cell in next_row
                        if cell is not None and str(cell).strip() != ""
                    )

                    # If the next row has significantly fewer populated cells or is text-heavy,
                    # consider this the start of footer
                    if (
                        next_row_populated < avg_populated_cells * 0.5
                        or sum(
                            1
                            for cell in next_row
                            if cell is not None
                            and isinstance(cell, str)
                            and len(str(cell)) > 20
                        )
                        > 0
                    ):
                        footer_start_idx = i
                        break
                continue

            # Count populated cells
            populated_cells = sum(
                1
                for cell in current_row
                if cell is not None and str(cell).strip() != ""
            )

            # Check for footer indicators
            footer_indicators = [
                "total",
                "sum",
                "average",
                "mean",
                "source",
                "note",
                "footnote",
            ]
            has_footer_text = any(
                cell is not None
                and isinstance(cell, str)
                and any(
                    indicator in str(cell).lower() for indicator in footer_indicators
                )
                for cell in current_row
            )

            # Check for the summary row
            looks_like_summary = has_footer_text and populated_cells <= table_width

            # Check for notes or sources (often longer text spanning multiple columns)
            long_text_cells = sum(
                1
                for cell in current_row
                if cell is not None and isinstance(cell, str) and len(str(cell)) > 50
            )

            # If this looks like the start of the footer, mark it
            if (
                (populated_cells < avg_populated_cells * 0.7 and i > header_row_idx + 3)
                or looks_like_summary
                or long_text_cells > 0
            ):
                footer_start_idx = i
                break

            # Check for inconsistent data types compared to data rows
            if i > header_row_idx + 3:
                data_type_mismatch = 0
                for j, cell in enumerate(current_row):
                    if j < len(header_row) and header_row[j] is not None:
                        # Get the most common data type for this column in previous rows
                        col_types = [
                            type(rows[idx][j]).__name__
                            for idx in range(header_row_idx + 1, i)
                            if idx < len(rows)
                            and j < len(rows[idx])
                            and rows[idx][j] is not None
                        ]
                        if col_types and cell is not None:
                            most_common_type = Counter(col_types).most_common(1)[0][0]
                            if type(cell).__name__ != most_common_type:
                                data_type_mismatch += 1

                # If many columns have type mismatches, this might be a footer row
                if data_type_mismatch > table_width * 0.5:
                    footer_start_idx = i
                    break

        return footer_start_idx

    @staticmethod
    def extract_metadata(rows: List[List[Any]]) -> Dict[str, Any]:
        metadata = {}

        for row in rows:
            if len(row) >= 2 and all(item is None for item in row[2:]):
                key, value = row[:2]
                if key is not None and value is not None:
                    metadata[str(key).strip().rstrip(":=").rstrip()] = str(
                        value
                    ).strip()

        return metadata

    @staticmethod
    def read_excel_properties(wb: Workbook) -> Dict[str, Any]:
        # Core properties from DocumentProperties
        core_props = wb.properties
        properties = {
            "title": core_props.title,
            "author": core_props.creator,
            "subject": core_props.subject,
            "description": core_props.description,
            "keywords": core_props.keywords,
            "category": core_props.category,
            "last_modified_by": core_props.lastModifiedBy,
            "created": core_props.created,
            "modified": core_props.modified,
            "status": core_props.contentStatus,
            "revision": core_props.revision,
            "version": core_props.version,
            "language": core_props.language,
            "identifier": core_props.identifier,
        }

        # Remove None values
        properties = {k: v for k, v in properties.items() if v is not None}

        # Assign custom properties if they exist
        if hasattr(wb, "custom_doc_props"):
            for prop in wb.custom_doc_props.props:
                if prop.value:
                    if prop.name in properties:
                        prop_name = f"custom.{prop.name}"
                    else:
                        prop_name = prop.name
                    properties[prop_name] = prop.value

        return properties
